"""
Logistics KPI Automation & Multi-Carrier Analytics Platform
Module: Executive Excel KPI Automation Engine (excel_builder.py)

Extracts aggregated and granular data from DuckDB OLAP engine and compiles a
pristine, C-Level Executive KPI Dashboard in Excel (.xlsx) with dynamic formatting,
KPI cards, conditional formatting rules, and executive summaries.
"""

import os
import sys
import time
import argparse
from datetime import datetime

# Ensure project root is in sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

try:
    from src.dimensional_model import DimensionalModelEngine
except ImportError:
    from dimensional_model import DimensionalModelEngine


# -----------------------------------------------------------------------------
# PALETTE & DESIGN SYSTEM (C-Level Executive Slate & Blue)
# -----------------------------------------------------------------------------
FONT_FAMILY = "Segoe UI"
COLOR_PRIMARY_DARK = "0F172A"      # Deep Navy
COLOR_HEADER_BG = "1E293B"         # Slate Dark
COLOR_CARD_BG = "F1F5F9"           # Crisp Light Slate
COLOR_CARD_BORDER = "CBD5E1"       # Subtle Border
COLOR_ACCENT_BLUE = "2563EB"       # Royal Blue
COLOR_SUCCESS_GREEN = "059669"     # Emerald Green
COLOR_WARNING_AMBER = "D97706"     # Amber
COLOR_DANGER_RED = "DC2626"        # Crimson Red
COLOR_ZEBRA_ROW = "F8FAFC"         # Soft Alternating Fill


class ExecutiveExcelBuilder:
    """
    Automated C-Level Excel Report Generator.
    Builds structured, formatted, and auditable spreadsheets directly from DuckDB.
    """

    def __init__(self, output_path: str = "dist/Executive_Logistics_KPI_Dashboard.xlsx"):
        self.output_path = output_path
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.engine = DimensionalModelEngine()
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

    def _apply_card_style(self, ws, start_col: int, start_row: int, title: str, value: str, subtext: str, color_accent: str):
        """Builds a polished executive KPI Card block (3 rows x 2 cols)."""
        card_fill = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")
        border_side = Side(style="thin", color=COLOR_CARD_BORDER)
        border = Border(top=border_side, left=border_side, right=border_side, bottom=border_side)
        
        # Merge cells for card layout
        c1 = get_column_letter(start_col)
        c2 = get_column_letter(start_col + 1)
        
        ws.merge_cells(f"{c1}{start_row}:{c2}{start_row}")
        ws.merge_cells(f"{c1}{start_row+1}:{c2}{start_row+1}")
        ws.merge_cells(f"{c1}{start_row+2}:{c2}{start_row+2}")

        # Top Label
        top_cell = ws[f"{c1}{start_row}"]
        top_cell.value = title.upper()
        top_cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color="64748B")
        top_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Main Metric Value
        val_cell = ws[f"{c1}{start_row+1}"]
        val_cell.value = value
        val_cell.font = Font(name=FONT_FAMILY, size=18, bold=True, color=color_accent)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Subtext / Delta
        sub_cell = ws[f"{c1}{start_row+2}"]
        sub_cell.value = subtext
        sub_cell.font = Font(name=FONT_FAMILY, size=8, italic=True, color="475569")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in range(start_row, start_row + 3):
            for c in range(start_col, start_col + 2):
                cell = ws.cell(row=r, column=c)
                cell.fill = card_fill
                cell.border = border

    def build_executive_summary_sheet(self):
        """Constructs the high-impact Executive Overview sheet."""
        ws = self.wb.create_sheet(title="Executive Summary")
        ws.views.sheetView[0].showGridLines = True

        # Sheet Banner Header
        ws.merge_cells("A1:I2")
        header = ws["A1"]
        header.value = "SKYDROPX / FRENET — EXECUTIVE LOGISTICS & CARRIER PERFORMANCE DASHBOARD"
        header.font = Font(name=FONT_FAMILY, size=14, bold=True, color="FFFFFF")
        header.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        header.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Retrieve Core Metrics from DuckDB
        summary_query = """
            SELECT 
                COUNT(*) AS total_shipments,
                SUM(gross_revenue) AS total_gmv,
                SUM(billed_carrier_cost) AS total_cost,
                SUM(net_margin) AS total_margin,
                ROUND((SUM(net_margin) / SUM(gross_revenue)) * 100, 2) AS margin_pct,
                ROUND((1.0 - (SUM(is_sla_breached) * 1.0 / COUNT(*))) * 100, 2) AS otd_rate,
                SUM(carrier_cost_variance) AS cost_variance,
                SUM(is_weight_discrepancy) AS weight_discrepancies
            FROM fact_shipments;
        """
        metrics = self.engine.conn.execute(summary_query).fetchone()
        tot_shipments, tot_gmv, tot_cost, tot_margin, margin_pct, otd_pct, cost_var, wt_discrep = metrics

        # Render 4 Executive Cards
        self._apply_card_style(ws, start_col=1, start_row=4, title="Total Shipments", 
                               value=f"{tot_shipments:,}", subtext="Aggregated Platform Orders", color_accent=COLOR_PRIMARY_DARK)
        
        self._apply_card_style(ws, start_col=3, start_row=4, title="Gross Revenue (GMV)", 
                               value=f"${tot_gmv:,.2f}", subtext="Merchant Shipping Billing", color_accent=COLOR_ACCENT_BLUE)
        
        self._apply_card_style(ws, start_col=5, start_row=4, title="Net Operating Margin", 
                               value=f"${tot_margin:,.2f} ({margin_pct:.1f}%)", subtext="Platform Gross Profit", color_accent=COLOR_SUCCESS_GREEN)
        
        self._apply_card_style(ws, start_col=7, start_row=4, title="On-Time Delivery (OTD)", 
                               value=f"{otd_pct:.1f}%", subtext="Target: >= 92.0% SLA", 
                               color_accent=COLOR_SUCCESS_GREEN if otd_pct >= 90 else COLOR_DANGER_RED)

        # Carrier Scorecard Table
        ws["A8"].value = "CARRIER OPERATIONAL & FINANCIAL SCORECARD"
        ws["A8"].font = Font(name=FONT_FAMILY, size=11, bold=True, color=COLOR_PRIMARY_DARK)
        
        headers = [
            "Carrier ID", "Carrier Name", "Service Level", "Tier", "Volume",
            "Avg Transit (Days)", "SLA Breaches", "OTD Rate %", "Gross GMV ($)", 
            "Carrier Cost ($)", "Net Margin ($)", "Margin %", "Cost Leakage ($)"
        ]
        
        # Write Table Headers
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col_idx, value=h)
            cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_PRIMARY_DARK, end_color=COLOR_PRIMARY_DARK, fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if "ID" in h or "%" in h else "left", vertical="center")

        df_scorecard = self.engine.get_carrier_scorecard()
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        for row_idx, row in df_scorecard.iterrows():
            current_row = 10 + row_idx
            is_zebra = (row_idx % 2 == 1)
            row_fill = PatternFill(start_color=COLOR_ZEBRA_ROW, end_color=COLOR_ZEBRA_ROW, fill_type="solid") if is_zebra else None

            ws.cell(row=current_row, column=1, value=row["carrier_id"]).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2, value=row["carrier_name"])
            ws.cell(row=current_row, column=3, value=row["service_level"])
            ws.cell(row=current_row, column=4, value=row["contract_tier"])
            
            c_vol = ws.cell(row=current_row, column=5, value=int(row["total_shipments"]))
            c_vol.number_format = "#,##0"
            
            c_tra = ws.cell(row=current_row, column=6, value=float(row["avg_transit_days"]))
            c_tra.number_format = "0.00"
            
            c_bre = ws.cell(row=current_row, column=7, value=int(row["total_sla_breaches"]))
            c_bre.number_format = "#,##0"
            
            c_otd = ws.cell(row=current_row, column=8, value=float(row["otd_rate_pct"]) / 100.0)
            c_otd.number_format = "0.0%"
            
            c_gmv = ws.cell(row=current_row, column=9, value=float(row["total_gross_revenue"]))
            c_gmv.number_format = "$#,##0.00"
            
            c_cst = ws.cell(row=current_row, column=10, value=float(row["total_carrier_cost"]))
            c_cst.number_format = "$#,##0.00"
            
            c_mar = ws.cell(row=current_row, column=11, value=float(row["total_net_margin"]))
            c_mar.number_format = "$#,##0.00"
            
            c_mpc = ws.cell(row=current_row, column=12, value=float(row["avg_margin_pct"]) / 100.0)
            c_mpc.number_format = "0.0%"
            
            c_leak = ws.cell(row=current_row, column=13, value=float(row["total_cost_leakage"]))
            c_leak.number_format = "$#,##0.00"

            for c in range(1, 14):
                cell = ws.cell(row=current_row, column=c)
                cell.font = Font(name=FONT_FAMILY, size=9)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill

        # Conditional Formatting on OTD Rate Column (H10:H16)
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ws.conditional_formatting.add("H10:H20", CellIsRule(operator="greaterThanOrEqual", formula=["0.90"], fill=green_fill))
        ws.conditional_formatting.add("H10:H20", CellIsRule(operator="lessThan", formula=["0.85"], fill=red_fill))

        # Autofit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def build_regional_and_merchants_sheets(self):
        """Builds the Regional Matrix and Merchant Performance sheets."""
        # 1. Regional SLA Matrix Sheet
        ws_reg = self.wb.create_sheet(title="Regional SLA Matrix")
        ws_reg.views.sheetView[0].showGridLines = True
        
        ws_reg["A1"].value = "REGIONAL LOGISTICS ROUTE & SLA PERFORMANCE MATRIX"
        ws_reg["A1"].font = Font(name=FONT_FAMILY, size=12, bold=True, color=COLOR_PRIMARY_DARK)
        
        df_reg = self.engine.get_regional_matrix()
        reg_headers = ["Destination Region", "Zone Classification", "Volume", "Avg Transit Days", "OTD Rate %", "Regional GMV ($)", "Net Margin ($)"]
        
        for c_idx, h in enumerate(reg_headers, 1):
            cell = ws_reg.cell(row=3, column=c_idx, value=h)
            cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center" if "%" in h else "left")

        for r_idx, r in df_reg.iterrows():
            curr = 4 + r_idx
            ws_reg.cell(row=curr, column=1, value=r["destination_region"])
            ws_reg.cell(row=curr, column=2, value=r["zone_type"])
            ws_reg.cell(row=curr, column=3, value=int(r["shipment_volume"])).number_format = "#,##0"
            ws_reg.cell(row=curr, column=4, value=float(r["avg_transit_days"])).number_format = "0.00"
            ws_reg.cell(row=curr, column=5, value=float(r["otd_pct"]) / 100.0).number_format = "0.0%"
            ws_reg.cell(row=curr, column=6, value=float(r["regional_gmv"])).number_format = "$#,##0.00"
            ws_reg.cell(row=curr, column=7, value=float(r["regional_net_margin"])).number_format = "$#,##0.00"

        for col in ws_reg.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_reg.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 14)

        # 2. Merchant Performance Sheet
        ws_mch = self.wb.create_sheet(title="Merchant Insights")
        ws_mch.views.sheetView[0].showGridLines = True
        
        ws_mch["A1"].value = "MERCHANT VOLUME, MARGIN & AUDIT DISCREPANCIES"
        ws_mch["A1"].font = Font(name=FONT_FAMILY, size=12, bold=True, color=COLOR_PRIMARY_DARK)
        
        df_mch = self.engine.get_merchant_performance()
        mch_headers = ["Merchant ID", "Merchant Name", "Tier", "Industry Category", "Shipments", "Gross GMV ($)", "Net Margin ($)", "Margin %", "Weight Discrepancies"]
        
        for c_idx, h in enumerate(mch_headers, 1):
            cell = ws_mch.cell(row=3, column=c_idx, value=h)
            cell.font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")

        for r_idx, r in df_mch.iterrows():
            curr = 4 + r_idx
            ws_mch.cell(row=curr, column=1, value=r["merchant_id"]).alignment = Alignment(horizontal="center")
            ws_mch.cell(row=curr, column=2, value=r["merchant_name"])
            ws_mch.cell(row=curr, column=3, value=r["merchant_tier"])
            ws_mch.cell(row=curr, column=4, value=r["industry_category"])
            ws_mch.cell(row=curr, column=5, value=int(r["total_shipments"])).number_format = "#,##0"
            ws_mch.cell(row=curr, column=6, value=float(r["total_gross_revenue"])).number_format = "$#,##0.00"
            ws_mch.cell(row=curr, column=7, value=float(r["total_net_margin"])).number_format = "$#,##0.00"
            ws_mch.cell(row=curr, column=8, value=float(r["margin_pct"]) / 100.0).number_format = "0.0%"
            ws_mch.cell(row=curr, column=9, value=int(r["audited_discrepancies"])).number_format = "#,##0"

        for col in ws_mch.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_mch.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 14)

    def build_raw_audit_feed_sheet(self, limit: int = 150):
        """Builds a curated sample drill-down feed of recent shipments."""
        ws = self.wb.create_sheet(title="Shipment Audit Feed")
        ws.views.sheetView[0].showGridLines = True

        query = f"""
            SELECT 
                shipment_id, tracking_number, merchant_id, carrier_id,
                order_timestamp, delivery_status, declared_weight_kg, billed_weight_kg,
                quoted_shipping_fee, billed_carrier_cost, gross_revenue, net_margin,
                transit_time_days, promised_sla_days, is_sla_breached
            FROM fact_shipments
            ORDER BY order_timestamp DESC
            LIMIT {limit}
        """
        df_feed = self.engine.conn.execute(query).df()
        
        headers = list(df_feed.columns)
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = Font(name=FONT_FAMILY, size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")

        for r_idx, row in df_feed.iterrows():
            curr = 2 + r_idx
            for c_idx, col_name in enumerate(headers, 1):
                val = row[col_name]
                cell = ws.cell(row=curr, column=c_idx, value=val)
                cell.font = Font(name=FONT_FAMILY, size=8)
                if "fee" in col_name or "cost" in col_name or "revenue" in col_name or "margin" in col_name:
                    cell.number_format = "$#,##0.00"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 12)

    def generate_workbook(self):
        """Generates the full executive dashboard workbook and saves to disk."""
        print(f"[Excel Builder] Compiling C-Level Executive Dashboard to {self.output_path}...")
        self.build_executive_summary_sheet()
        self.build_regional_and_merchants_sheets()
        self.build_raw_audit_feed_sheet()
        self.wb.save(self.output_path)
        print(f"[Excel Builder] Workbook successfully generated at: {self.output_path}")


def main():
    print("=" * 80)
    print(">> C-LEVEL EXECUTIVE EXCEL DASHBOARD AUTOMATION ENGINE")
    print("=" * 80)
    start_time = time.time()
    
    builder = ExecutiveExcelBuilder()
    builder.generate_workbook()
    
    elapsed = (time.time() - start_time) * 1000
    print(f">> [Excel Builder] Completed in {elapsed:.2f} ms")


if __name__ == "__main__":
    main()
